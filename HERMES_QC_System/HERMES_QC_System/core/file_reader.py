"""
HERMES QC System - File Reader Module
Reads PDF and Excel files from OneDrive QC_Report folder
"""

import os
import json
from pathlib import Path
from datetime import datetime, timedelta
import pypdf
from openpyxl import load_workbook


class QCFileReader:
    """Read PDF and Excel inspection reports from a folder"""
    
    def __init__(self, folder_path):
        self.folder_path = Path(folder_path)
        if not self.folder_path.exists():
            raise FileNotFoundError(f"Folder not found: {folder_path}")
    
    def get_files_by_date(self, target_date=None, days_back=0):
        """
        Get all PDF/Excel files modified on a specific date
        
        Args:
            target_date: datetime.date object (default: today)
            days_back: How many days back to include (0 = only target_date)
        
        Returns:
            List of file paths
        """
        if target_date is None:
            target_date = datetime.now().date()
        
        start_date = target_date - timedelta(days=days_back)
        end_date = target_date + timedelta(days=1)
        
        valid_extensions = {'.pdf', '.xlsx', '.xls', '.xlsm'}
        matched_files = []
        
        for file_path in self.folder_path.iterdir():
            if file_path.is_file() and file_path.suffix.lower() in valid_extensions:
                mtime = datetime.fromtimestamp(file_path.stat().st_mtime).date()
                if start_date <= mtime < end_date:
                    matched_files.append(file_path)
        
        return sorted(matched_files, key=lambda f: f.stat().st_mtime)
    
    def read_pdf(self, file_path):
        """Extract text content from a PDF file"""
        try:
            reader = pypdf.PdfReader(str(file_path))
            text_content = []
            for page_num, page in enumerate(reader.pages, start=1):
                page_text = page.extract_text()
                text_content.append(f"--- Page {page_num} ---\n{page_text}")
            
            return {
                "filename": file_path.name,
                "pages": len(reader.pages),
                "content": "\n\n".join(text_content),
                "file_type": "PDF",
                "status": "success"
            }
        except Exception as e:
            return {
                "filename": file_path.name,
                "content": "",
                "file_type": "PDF",
                "status": "error",
                "error": str(e)
            }
    
    def read_excel(self, file_path):
        """Extract data from Excel file"""
        try:
            wb = load_workbook(str(file_path), data_only=True)
            all_content = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_data = [f"--- Sheet: {sheet_name} ---"]
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(c) if c is not None else "" for c in row)
                    if row_text.strip():
                        sheet_data.append(row_text)
                
                all_content.append("\n".join(sheet_data))
            
            return {
                "filename": file_path.name,
                "sheets": len(wb.sheetnames),
                "content": "\n\n".join(all_content),
                "file_type": "Excel",
                "status": "success"
            }
        except Exception as e:
            return {
                "filename": file_path.name,
                "content": "",
                "file_type": "Excel",
                "status": "error",
                "error": str(e)
            }
    
    def read_file(self, file_path):
        """Auto-detect and read file"""
        ext = Path(file_path).suffix.lower()
        if ext == '.pdf':
            return self.read_pdf(file_path)
        elif ext in {'.xlsx', '.xls', '.xlsm'}:
            return self.read_excel(file_path)
        else:
            return {"filename": Path(file_path).name, "status": "unsupported"}
    
    def read_all_today(self):
        """Read all today's files - main entry point"""
        files = self.get_files_by_date()
        results = []
        for f in files:
            print(f"📄 Reading: {f.name}")
            results.append(self.read_file(f))
        return results


if __name__ == "__main__":
    # Test
    config_path = Path(__file__).parent.parent / "config" / "config.json"
    with open(config_path) as f:
        config = json.load(f)
    
    reader = QCFileReader(config["paths"]["onedrive_qc_folder"])
    results = reader.read_all_today()
    print(f"\n✅ Total files read: {len(results)}")
