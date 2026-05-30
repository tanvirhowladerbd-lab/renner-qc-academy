"""
HERMES QC System - Excel Summary Generator
Generates consolidated single-table Excel matching Tanvir's approved format
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path


class ExcelSummaryGenerator:
    """Generate the consolidated QC summary Excel"""
    
    # Styles
    THIN = Side(border_style="thin", color="000000")
    BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    
    COLOR_HEADER = "1F4E78"
    COLOR_SUBHEADER = "2E75B6"
    COLOR_RED = "FFC7CE"
    COLOR_YELLOW = "FFEB9C"
    COLOR_GREEN = "C6EFCE"
    COLOR_GRAY = "F2F2F2"
    COLOR_BLUE = "DDEBF7"
    
    FONT_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    FONT_BODY = Font(name="Arial", size=10, color="000000")
    FONT_BOLD = Font(name="Arial", size=10, bold=True, color="000000")
    FONT_STATUS_RED = Font(name="Arial", size=10, bold=True, color="9C0006")
    FONT_STATUS_YELLOW = Font(name="Arial", size=10, bold=True, color="9C5700")
    FONT_STATUS_GREEN = Font(name="Arial", size=10, bold=True, color="006100")
    
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    def __init__(self, output_folder, recipient_email=""):
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(parents=True, exist_ok=True)
        self.recipient_email = recipient_email
    
    @property
    def fills(self):
        return {
            "header": PatternFill("solid", start_color=self.COLOR_HEADER),
            "subheader": PatternFill("solid", start_color=self.COLOR_SUBHEADER),
            "red": PatternFill("solid", start_color=self.COLOR_RED),
            "yellow": PatternFill("solid", start_color=self.COLOR_YELLOW),
            "green": PatternFill("solid", start_color=self.COLOR_GREEN),
            "gray": PatternFill("solid", start_color=self.COLOR_GRAY),
            "blue": PatternFill("solid", start_color=self.COLOR_BLUE),
        }
    
    def generate(self, reports_data, run_date=None):
        """
        Generate Excel summary
        
        Args:
            reports_data: List of dicts (output from ClaudeAnalyzer)
            run_date: datetime object (default: now)
        
        Returns: Path to generated Excel file
        """
        if run_date is None:
            run_date = datetime.now()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily QC Summary"
        fills = self.fills
        
        # ========== TITLE ==========
        ws.merge_cells("A1:K1")
        ws["A1"] = "MOVIMODA - QC DAILY CONSOLIDATED SUMMARY"
        ws["A1"].font = self.FONT_TITLE
        ws["A1"].fill = fills["header"]
        ws["A1"].alignment = self.ALIGN_CENTER
        ws.row_dimensions[1].height = 30
        
        # ========== SUBTITLE ==========
        ws.merge_cells("A2:K2")
        ws["A2"] = (f"Date: {run_date.strftime('%d-%b-%Y')}  |  "
                    f"Generated: {run_date.strftime('%H:%M')}  |  "
                    f"Recipient: {self.recipient_email}")
        ws["A2"].font = Font(name="Arial", size=10, italic=True, color="FFFFFF")
        ws["A2"].fill = fills["subheader"]
        ws["A2"].alignment = self.ALIGN_CENTER
        ws.row_dimensions[2].height = 22
        
        # ========== STATS BAR ==========
        total = len(reports_data)
        pass_count = sum(1 for r in reports_data if r.get("final_status") == "PASS")
        hold_count = sum(1 for r in reports_data if r.get("final_status") == "HOLD")
        reject_count = sum(1 for r in reports_data if r.get("final_status") == "REJECT")
        other_count = total - pass_count - hold_count - reject_count
        
        stats_text = (f"📊 Today: {total} Reports  |  "
                     f"✅ PASS: {pass_count}  |  "
                     f"⚠️ HOLD: {hold_count}  |  "
                     f"🔴 REJECT: {reject_count}  |  "
                     f"📋 Others: {other_count}")
        
        ws.merge_cells("A3:K3")
        ws["A3"] = stats_text
        ws["A3"].font = Font(name="Arial", size=11, bold=True)
        ws["A3"].fill = fills["blue"]
        ws["A3"].alignment = self.ALIGN_CENTER
        ws.row_dimensions[3].height = 26
        
        ws.row_dimensions[4].height = 8
        
        # ========== TABLE HEADERS ==========
        headers = ["SL", "Date", "Vendor", "Style", "Job Order/Lot", "Lot Qty",
                  "Inspection Type", "Status", "Issue / Defects Found",
                  "Vendor Action Required", "CAP Status"]
        
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = self.FONT_HEADER
            cell.fill = fills["header"]
            cell.alignment = self.ALIGN_CENTER
            cell.border = self.BORDER_ALL
        ws.row_dimensions[5].height = 32
        
        # ========== DATA ROWS ==========
        # Sort: REJECT first, then HOLD, then PASS, then others
        status_order = {"REJECT": 0, "HOLD": 1, "PP STAGE": 2, "PASS": 3}
        sorted_reports = sorted(reports_data,
                              key=lambda r: status_order.get(r.get("final_status", "PASS"), 4))
        
        for idx, r in enumerate(sorted_reports, start=1):
            row_num = 5 + idx
            
            row_data = [
                idx,
                r.get("inspection_date", run_date.strftime("%d-%b-%y")),
                r.get("vendor", "N/A"),
                r.get("style", "N/A"),
                r.get("job_order", "N/A"),
                r.get("lot_qty", "N/A"),
                r.get("inspection_type", "N/A"),
                r.get("final_status", "N/A"),
                r.get("main_issue", "") + 
                    (f"; Defects: {r.get('defect_details', '')}" 
                     if r.get('defect_details') and r.get('defect_details') != 'N/A' else ""),
                r.get("vendor_action_required", "N/A"),
                "Pending"  # Default CAP status
            ]
            
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = self.BORDER_ALL
                cell.font = self.FONT_BODY
                cell.alignment = self.ALIGN_LEFT if col in [9, 10] else self.ALIGN_CENTER
            
            # Color status cell
            status_cell = ws.cell(row=row_num, column=8)
            status = r.get("final_status", "")
            if status == "REJECT":
                status_cell.fill = fills["red"]
                status_cell.font = self.FONT_STATUS_RED
            elif status == "HOLD":
                status_cell.fill = fills["yellow"]
                status_cell.font = self.FONT_STATUS_YELLOW
            elif status == "PASS":
                status_cell.fill = fills["green"]
                status_cell.font = self.FONT_STATUS_GREEN
            else:
                status_cell.fill = fills["gray"]
                status_cell.font = self.FONT_BOLD
            
            # CAP status (default Pending = Red)
            cap_cell = ws.cell(row=row_num, column=11)
            cap_cell.fill = fills["red"]
            cap_cell.font = self.FONT_STATUS_RED
            
            # Row height
            content_len = len(str(row_data[8])) + len(str(row_data[9]))
            if content_len > 200:
                ws.row_dimensions[row_num].height = 75
            elif content_len > 100:
                ws.row_dimensions[row_num].height = 55
            else:
                ws.row_dimensions[row_num].height = 40
        
        # Freeze header
        ws.freeze_panes = "A6"
        
        # ========== COLUMN WIDTHS ==========
        col_widths = {
            "A": 5, "B": 12, "C": 24, "D": 14, "E": 14, "F": 10,
            "G": 18, "H": 12, "I": 50, "J": 45, "K": 14
        }
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w
        
        # ========== FOOTER LEGEND ==========
        footer_row = 5 + len(sorted_reports) + 2
        
        ws.cell(row=footer_row, column=1, value="LEGEND:").font = self.FONT_BOLD
        ws.merge_cells(start_row=footer_row, start_column=1,
                      end_row=footer_row, end_column=11)
        ws.cell(row=footer_row, column=1).fill = fills["gray"]
        
        legends = [
            ("🟢 PASS", "Goods acceptable, ready for shipment", "green"),
            ("🟡 HOLD", "Goods on hold, rework/CAP required before shipment", "yellow"),
            ("🔴 REJECT", "Goods rejected, full re-inspection mandatory", "red"),
        ]
        
        for i, (lbl, desc, color) in enumerate(legends):
            r = footer_row + 1 + i
            ws.cell(row=r, column=1, value=lbl).font = self.FONT_BOLD
            ws.cell(row=r, column=1).fill = fills[color]
            ws.cell(row=r, column=1).alignment = self.ALIGN_CENTER
            ws.cell(row=r, column=1).border = self.BORDER_ALL
            
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
            ws.cell(row=r, column=2, value=desc).font = self.FONT_BODY
            ws.cell(row=r, column=2).alignment = self.ALIGN_LEFT
            ws.cell(row=r, column=2).border = self.BORDER_ALL
        
        # Final note
        note_row = footer_row + 1 + len(legends) + 1
        ws.merge_cells(start_row=note_row, start_column=1,
                      end_row=note_row, end_column=11)
        ws.cell(row=note_row, column=1, value=
            "⚠️ NOTE: Review HOLD and REJECT items first. Send CAP follow-up to vendors. "
            "Auto-generated by HERMES from OneDrive/QC_Report folder."
        ).font = Font(name="Arial", size=9, italic=True, color="9C0006")
        ws.cell(row=note_row, column=1).fill = fills["yellow"]
        ws.cell(row=note_row, column=1).alignment = self.ALIGN_LEFT
        
        # Save
        filename = f"QC_Daily_Summary_{run_date.strftime('%Y-%m-%d')}.xlsx"
        output_path = self.output_folder / filename
        wb.save(str(output_path))
        
        return str(output_path)


if __name__ == "__main__":
    # Test with mock data
    mock_data = [
        {
            "vendor": "CONCEPT KNITTING LTD",
            "style": "3TNHL108D",
            "job_order": "26S_1095",
            "lot_qty": 1138,
            "final_status": "HOLD",
            "main_issue": "Individual Size Short 65.3%",
            "vendor_action_required": "100% measurement re-check, RCA, Rework plan",
            "inspection_type": "Final Random",
            "inspection_date": "13-May-26",
        }
    ]
    
    gen = ExcelSummaryGenerator("/tmp/test_output")
    path = gen.generate(mock_data)
    print(f"✅ Generated: {path}")
